"""
    This file was used to search 1nn and 2nn of a atom.
"""
import atoms
import numpy as np
import copy
from openpyxl import Workbook

class atompair:
    """
    A class representing a pair of atoms, typically used for neighbor analysis.

    Attributes:
        oatom (atoms.atom): The central or reference atom.
        patom (atoms.atom): The neighboring atom in its original coordinates.
        patomtrasf (atoms.atom): The same neighboring atom with coordinates adjusted
                                 to account for periodic boundary conditions (minimum image).
        distance (np.float64): The distance between patom and oatom, computed with
                               periodic boundary conditions taken into account.
    """

    def __init__(self, oatom: atoms.atom, patom: atoms.atom, patomtrasf: atoms.atom, distance: np.float64):
        """
        Initialize an atom pair.

        Parameters:
            oatom (atoms.atom): The core/center atom.
            patom (atoms.atom): The other atom (neighbor) in its original coordinates.
            patomtrasf (atoms.atom): The same as patom, but with coordinates transformed
                                     according to the periodic boundary conditions (minimum image).
            distance (np.float64): The distance between patom and oatom, calculated with
                                   periodic boundary conditions.
        """
        self.oatom = oatom
        self.patom = patom
        self.patomtrasf = patomtrasf
        self.distance = distance

def finding(atomlist: list[atoms.atom], vectors: list[np.float64]):
    """
    Find and classify nearest neighbor pairs for each atom in the system.

    For each central atom (oatom), this function identifies all other atoms (patom)
    and calculates their minimum-image distances. The neighbors are then sorted by
    distance and classified into first, second, third, and fourth nearest neighbor
    shells based on fixed cutoffs (the 8th, 14th, 26th, and 50th closest distances
    for a bcc-like structure, assuming 8+6+12+24 = 50 neighbors in the first four shells).

    Parameters:
        atomlist (list[atoms.atom]): List of atom objects (must have .coord attribute).
        vectors (list[np.float64]): List of three box vectors (v1, v2, v3) as returned by boxGET().
                                    Assumed to be aligned with Cartesian axes.

    Returns:
        tuple: Four lists, each containing for every central atom a list of atompair objects
               corresponding to the first, second, third, and fourth nearest neighbor shells.
               The order of central atoms follows the order in atomlist.
    """
    # Initialize result containers
    listofatompairsof1nn: list = []  # First nearest neighbors for each center
    listofatompairsof2nn: list = []  # Second nearest neighbors
    listofatompairsof3nn: list = []  # Third nearest neighbors
    listofatompairsof4nn: list = []  # Fourth nearest neighbors

    for oatom in atomlist:  # Loop over each atom as the center
        atompairs = []       # Temporary list to store all pairs for this center
        boundaries = []      # To store half-box boundaries (not used correctly, but kept as is)
        distances = []       # List of distances to all other atoms (after PBC correction)

        for patom in atomlist:
            if oatom == patom:  # Skip self (comparison might rely on coordinates; ensure atom objects are comparable)
                continue
            else:
                # Recompute boundaries for each axis (inefficient, but kept as original)
                for axis in np.arange(3):
                    boundaries.append([oatom.coord + vectors[axis]/2, oatom.coord - vectors[axis]/2])

                # Create a transformed copy of patom with coordinates adjusted for periodic boundary conditions
                patomtrasf = copy.deepcopy(patom)
                for axis in np.arange(3):
                    # Apply minimum image convention using the stored boundaries
                    if patomtrasf.coord[axis] > boundaries[axis][0][axis]:
                        patomtrasf.coord[axis] -= vectors[axis][axis]
                    elif patomtrasf.coord[axis] < boundaries[axis][1][axis]:
                        patomtrasf.coord[axis] += vectors[axis][axis]

                # Create an atompair object for this neighbor
                dist = atoms.distenceCalc(oatom.coord, patomtrasf.coord)
                atompairs.append(atompair(oatom=oatom, patom=patom, patomtrasf=patomtrasf, distance=dist))
                distances.append(dist)

        # Make a deep copy of atompairs before filtering
        atompairs_cp = copy.deepcopy(atompairs)
        atompairs = []  # Reset for collecting filtered pairs

        # Sort distances to determine shell cutoffs
        distances_cp = sorted(distances)

        # Boolean masks for each shell based on indices (assuming at least 50 atoms in system)
        # For a bcc structure: 1nn = 8, 2nn = 6, 3nn = 12, 4nn = 24 → total 50.
        # Cutoffs: distances <= 8th (index 7) → 1nn; 9th to 14th (indices 8-13) → 2nn;
        #          15th to 26th (indices 14-25) → 3nn; 27th to 50th (indices 26-49) → 4nn.
        locationsof1nn = distances <= distances_cp[7]
        locationsof2nn = ((distances <= distances_cp[13]) & (distances > distances_cp[7]))
        locationsof3nn = ((distances <= distances_cp[25]) & (distances > distances_cp[13]))
        locationsof4nn = ((distances <= distances_cp[49]) & (distances > distances_cp[25]))

        # Collect first nearest neighbors (max 8)
        iflg = 0
        for location in locationsof1nn:
            if location:
                atompairs.append(atompairs_cp[iflg])
            iflg += 1
            if len(atompairs) > 8:
                break
        listofatompairsof1nn.append(atompairs)

        # Collect second nearest neighbors (max 6)
        atompairs = []
        iflg = 0
        for location in locationsof2nn:
            if location:
                atompairs.append(atompairs_cp[iflg])
            iflg += 1
            if len(atompairs) > 6:
                break
        listofatompairsof2nn.append(atompairs)

        # Collect third nearest neighbors (max 12)
        atompairs = []
        iflg = 0
        for location in locationsof3nn:
            if location:
                atompairs.append(atompairs_cp[iflg])
            iflg += 1
            if len(atompairs) > 12:
                break
        listofatompairsof3nn.append(atompairs)

        # Collect fourth nearest neighbors (max 24)
        atompairs = []
        iflg = 0
        for location in locationsof4nn:
            if location:
                atompairs.append(atompairs_cp[iflg])
            iflg += 1
            if len(atompairs) > 24:
                break
        listofatompairsof4nn.append(atompairs)

    return listofatompairsof1nn, listofatompairsof2nn, listofatompairsof3nn, listofatompairsof4nn

def documenting(listofatompairs: list[list[atompair]], listofelements):
    """
    :param listofatompairs: which gotten from function finding.
    :param setofelements: the constitute elements of the HEA
    :return:
    """
    storageofpairs = []
    dictofpairs = dict()
    for i in np.arange(len(listofelements)):
        for j in np.arange(i, len(listofelements)):
            dictofpairs['{}-{}'.format(listofelements[i], listofelements[j])] = []
    for atompairs in listofatompairs:
        for atompair in atompairs:
            if [str(atompair.oatom.coord), str(atompair.patom.coord)] in storageofpairs:
                continue
            else:
                try:
                    dictofpairs['{}-{}'.format(atompair.oatom.typ, atompair.patom.typ)].append(atompair.distance)
                except:
                    dictofpairs['{}-{}'.format(atompair.patom.typ, atompair.oatom.typ)].append(atompair.distance)
                storageofpairs.append([str(atompair.patom.coord), str(atompair.oatom.coord)])
    return dictofpairs

def fnnlocate(listofatompairsof1nn: list[list[atompair]]):
    # 三视图中的位置
    # 1——左后下
    # 2——左前下
    # 3——右前下
    # 4——右后下
    # 5——左后上
    # 6——左前上
    # 7——右前上
    # 8——右后上
    listoffnncluster = []
    for atompairs in listofatompairsof1nn:
        dic = {}
        dic['oatom'] = atompairs[0].oatom
        for atompair in atompairs:
            if atompair.patomtrasf.coord[0] < atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] < atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] < atompair.oatom.coord[2]:
                dic['patom1'] = atompair.patom
                # print(1)
            if atompair.patomtrasf.coord[0] > atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] < atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] < atompair.oatom.coord[2]:
                dic['patom2'] = atompair.patom
                # print(2)
            if atompair.patomtrasf.coord[0] > atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] > atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] < atompair.oatom.coord[2]:
                dic['patom3'] = atompair.patom
                # print(3)
            if atompair.patomtrasf.coord[0] < atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] > atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] < atompair.oatom.coord[2]:
                dic['patom4'] = atompair.patom
                # print(4)
            if atompair.patomtrasf.coord[0] < atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] < atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] > atompair.oatom.coord[2]:
                dic['patom5'] = atompair.patom
                # print(5)
            if atompair.patomtrasf.coord[0] > atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] < atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] > atompair.oatom.coord[2]:
                dic['patom6'] = atompair.patom
                # print(6)
            if atompair.patomtrasf.coord[0] > atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] > atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] > atompair.oatom.coord[2]:
                dic['patom7'] = atompair.patom
                # print(7)
            if atompair.patomtrasf.coord[0] < atompair.oatom.coord[0] and atompair.patomtrasf.coord[1] > atompair.oatom.coord[1] and atompair.patomtrasf.coord[2] > atompair.oatom.coord[2]:
                dic['patom8'] = atompair.patom
                # print(8)
        listoffnncluster.append(dic)
        # print('_________{}___________'.format(len(listoffnncluster)))
    return listoffnncluster

def snnlocate(listofatompairsof2nn):
    # 三视图中的位置
    # 9——前
    # 10——后
    # 11——左
    # 12——右
    # 13——上
    # 14——下
    listofsnncluster = []
    for atompairs in listofatompairsof2nn:
        dic = {}
        dic['oatom'] = atompairs[0].oatom
        for atompair in atompairs:
            if atompair.patomtrasf.coord[0] > atompair.oatom.coord[0] + 2.5:
                dic['patom9'] = atompair.patom
            if atompair.patomtrasf.coord[0] < atompair.oatom.coord[0] - 2.5:
                dic['patom10'] = atompair.patom
            if atompair.patomtrasf.coord[1] < atompair.oatom.coord[1] - 2.5:
                dic['patom11'] = atompair.patom
            if atompair.patomtrasf.coord[1] > atompair.oatom.coord[1] + 2.5:
                dic['patom12'] = atompair.patom
            if atompair.patomtrasf.coord[2] > atompair.oatom.coord[2] + 2.5:
                dic['patom13'] = atompair.patom
            if atompair.patomtrasf.coord[2] < atompair.oatom.coord[2] - 2.5:
                dic['patom14'] = atompair.patom
        listofsnncluster.append(dic)
    return listofsnncluster



if __name__ == '__main__':
    vectors = atoms.boxGET('./CONTCAR-rlx-s2')
    atomslist = atoms.atomsGET('./CONTCAR-rlx-s2', vectors)
    listofatompairsof1nn, listofatompairsof2nn, listofatompairsof3nn, listofatompairsof4nn = finding(atomslist, vectors)
    listoffnncluster: list[dict[str, atoms.atom]] = fnnlocate(listofatompairsof1nn)
    # print(listoffnncluster[0]['patom6'].typ)
    wb = Workbook()
    ws = wb.active
    for i in np.arange(128):
        ws['a{}'.format(i + 1)] = listoffnncluster[i]['patom1'].typ
        ws['b{}'.format(i + 1)] = listoffnncluster[i]['patom2'].typ
        ws['c{}'.format(i + 1)] = listoffnncluster[i]['patom3'].typ
        ws['d{}'.format(i + 1)] = listoffnncluster[i]['patom4'].typ
        ws['e{}'.format(i + 1)] = listoffnncluster[i]['patom5'].typ
        ws['f{}'.format(i + 1)] = listoffnncluster[i]['patom6'].typ
        ws['g{}'.format(i + 1)] = listoffnncluster[i]['patom7'].typ
        ws['h{}'.format(i + 1)] = listoffnncluster[i]['patom8'].typ
    wb.save('./output/fnnlocate.xlsx')
    listofsnncluster: list[dict[str, atoms.atom]] = snnlocate(listofatompairsof2nn)
    wb = Workbook()
    ws = wb.active
    for i in np.arange(128):
        ws['a{}'.format(i + 1)] = listofsnncluster[i]['patom9'].typ
        ws['b{}'.format(i + 1)] = listofsnncluster[i]['patom10'].typ
        ws['c{}'.format(i + 1)] = listofsnncluster[i]['patom11'].typ
        ws['d{}'.format(i + 1)] = listofsnncluster[i]['patom12'].typ
        ws['e{}'.format(i + 1)] = listofsnncluster[i]['patom13'].typ
        ws['f{}'.format(i + 1)] = listofsnncluster[i]['patom14'].typ
    wb.save('./output/snnlocate.xlsx')
    # nn1 = []
    # nn2 = []
    # nn3 = []
    # nn4 = []
    # wb = Workbook()
    # ws = wb.active
    # ws.title = 'Interatomic distance'
    # ws['a1'] = '1nn'
    # ws['b1'] = '2nn'
    # ws['c1'] = '3nn'
    # ws['d1'] = '4nn'
    # for atompairs in listofatompairsof1nn:
    #     for atompair in atompairs:
    #         nn1.append(atompair.distance)
    # for i in np.arange(len(nn1)):
    #     ws['a{}'.format(i+2)] = nn1[i]
    # for atompairs in listofatompairsof2nn:
    #     for atompair in atompairs:
    #         nn2.append(atompair.distance)
    # for i in np.arange(len(nn2)):
    #     ws['b{}'.format(i+2)] = nn2[i]
    # for atompairs in listofatompairsof3nn:
    #     for atompair in atompairs:
    #         nn3.append(atompair.distance)
    # for i in np.arange(len(nn3)):
    #     ws['c{}'.format(i+2)] = nn3[i]
    # for atompairs in listofatompairsof4nn:
    #     for atompair in atompairs:
    #         nn4.append(atompair.distance)
    # for i in np.arange(len(nn4)):
    #     ws['d{}'.format(i+2)] = nn4[i]
    # wb.save('InteratomicDistancesofNeighbors.xlsx')
