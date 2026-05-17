"""
要模拟在弛豫的过程中，|f|和|t|的夹角的变化，以及|f|随着|t|的变化
"""
import re
import math
from openpyxl import Workbook
import numpy as np

# 依据某一个标识("FREE ENERGIE OF THE ION-ELECTRON SYSTEM (eV)")获得每一个电子步计算完成之后所有原子的位置和原子上的力
wb = Workbook()
for k in [5, 6,  78, 79, 102, 103]:
    pathofoutcar = r'D:\gpc\FeCrV合金中空位的性质-第一篇文章\Vacancy formation energy-\Vac{}\OUTCAR-rlx'.format(k)

    with open(file=pathofoutcar, mode='rt', encoding='UTF-8') as f:
        list_lines = f.readlines()
    list_lines_strip = [lin.strip() for lin in list_lines]
    list_indexes = []
    for idx, lin in enumerate(list_lines_strip):
        if lin == 'FREE ENERGIE OF THE ION-ELECTRON SYSTEM (eV)':
            list_indexes.append(idx)

    # list_indexes中每一个idx都要对应两个列表
    list_forces = []
    list_positions = []
    # 获取所有的坐标和力
    for idx in list_indexes:
        forces = []
        positions = []
        for i in range(idx - 135, idx - 8):
            lin_split = re.split(r'\s+', list_lines_strip[i])
            lin_split_to_float = [float(e) for e in lin_split]
            positions += lin_split_to_float[0:3]
            forces += lin_split_to_float[3:6]
        list_forces.append(forces)
        list_positions.append(positions)

    # 根据第一帧的所有原子坐标，normalize所有原子坐标
    for i in range(1, len(list_positions)):
        for j in range(len(list_positions[i])):
            if list_positions[i][j] - list_positions[0][j] > 11.4876/2:
                list_positions[i][j] -= 11.486
            elif list_positions[i][j] - list_positions[0][j] < -11.486/2:
                list_positions[i][j] += 11.486

    # 计算位移
    list_displacements_interval = []
    list_displace_distence = []
    for position in list_positions[1:]:
        displacement = np.array(position) - np.array(list_positions[list_positions.index(position) - 1])
        # print(displacement)
        list_displacements_interval.append(displacement)
        list_displace_distence.append(np.linalg.norm(np.array(position) - np.array(list_positions[0])))
    # 计算每一步力和displacement的夹角以及功
    list_works = []
    list_angles = []
    for force, displacement in zip(list_forces[0:len(list_forces)-1], list_displacements_interval):
        work = 0
        for f, d in zip(force, displacement):
            work += f * d

        cos = work / (np.linalg.norm(np.array(force)) * np.linalg.norm(np.array(displacement)))
        list_works.append(work)
        list_angles.append(math.acos(cos) / math.pi * 180)
        print(work, math.acos(cos) / math.pi * 180)

    # list_displace_distence.insert(0, 0)
    # list_works.insert(0, 0)
    # list_angles.append(0)


    ws = wb.create_sheet('vac{}'.format(k))
    for j in range(len(list_works)):
        ws['a{}'.format(j + 1)] = list_displace_distence[j]
        ws['b{}'.format(j + 1)] = list_works[j]
        ws['c{}'.format(j + 1)] = list_angles[j]
wb.save('./output/relaxation_procedure.xlsx')











